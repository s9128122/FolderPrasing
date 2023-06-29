import os
import re
import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
import time

# GUI窗口
window = tk.Tk()
window.title("資料夾階層結構匯出程式")



# 資料夾路徑變數
folder_path = None

# 資料夾路徑選擇函數
def select_folder_path():
    global folder_path
    selected_folder_path = filedialog.askdirectory()
    #folder_path = selected_folder_path
    folder_path = selected_folder_path
    folder_label.config(text=f"資料夾路徑: {folder_path}")  # Update the folder path label

# 執行按鈕函數
def execute_program():
    # 獲取開始時間
    start_time = time.time()
    start_time_str = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(start_time))
    # 顯示執行結果
    text_output.delete("1.0", tk.END)  # 清空原有內容
    text_output.insert(tk.END, f"開始執行時間：{start_time_str}\n\n")        
    global folder_path

    try:
        current_directory = folder_path

        # 程式其餘部分...
        # 讀取文件夾和文件的階層結構
        folder_structure = {}
    
        # 讀取文件夾和文件的階層結構
        def read_folder_structure(folder_path, level=0):
            folder_name = os.path.basename(folder_path)
            folder_structure[folder_path] = {'name': folder_name, 'level': level, 'files': [], 'subfolders': []}
        
            for item in os.listdir(folder_path):
                item_path = os.path.join(folder_path, item)
                if os.path.isdir(item_path):
                    folder_structure[folder_path]['subfolders'].append(item_path)
                    read_folder_structure(item_path, level + 1)
                else:
                    folder_structure[folder_path]['files'].append(item)
        
        # 讀取文件夾和文件的階層結構
        def read_folder_structure(folder_path, level=0):
            folder_name = os.path.basename(folder_path)
            folder_structure[folder_path] = {'name': folder_name, 'level': level, 'files': [], 'subfolders': []}
        
            for item in os.listdir(folder_path):
                item_path = os.path.join(folder_path, item)
                if os.path.isdir(item_path):
                    folder_structure[folder_path]['subfolders'].append(item_path)
                    read_folder_structure(item_path, level + 1)
                else:
                    folder_structure[folder_path]['files'].append(item)
        
        # 讀取文件夾和文件的階層結構
        read_folder_structure(current_directory)

        # 統計資料夾數和檔數
        folder_count = 0
        file_count = 0

        for folder_info in folder_structure.values():
            folder_count += 1
            file_count += len(folder_info['files'])

        # 在文字方塊中插入統計結果
        text_output.insert(tk.END, f"資料夾總數：{folder_count}個\n")
        text_output.insert(tk.END, f"檔案總數：{file_count}個\n\n")
    

    
        def get_max_folder_level(folder_structure):
            max_level = 0
            for folder_info in folder_structure.values():
                level = folder_info['level']
                if level > max_level:
                    max_level = level
            return max_level
        
        # 獲得資料夾階層數量
        max_folder_level = get_max_folder_level(folder_structure)
        #print("資料夾階層數量：", max_folder_level)
        
    
        # 創建txt文件，記錄文件夾和文件的階層結構
        text_file_path = os.path.join(current_directory, '資料階層及檔案清單.txt')
        exclude_list=['資料階層及檔案清單.txt', '資料階層及檔案清單.xlsx', '列出資料名稱.py','列出資料名稱_fin.py','~$資料階層及檔案清單.xlsx']
        def write_folder_structure(folder_info, indent=0, file=None):
            folder_name = folder_info['name']
            if folder_name not in exclude_list:
                file.write(' ' * indent + folder_name + '/' + '\n')
                files = folder_info['files']
                for file_name in files:
                    if file_name not in exclude_list:
                        file.write(' ' * (indent + 4) + file_name + '\n')
    
                subfolders = folder_info['subfolders']
                for subfolder_path in subfolders:
                    subfolder_info = folder_structure[subfolder_path]
                    write_folder_structure(subfolder_info, indent + 4, file)

        # 輸出文件夾結構到txt
        with open(text_file_path, 'w', encoding='utf-8') as text_file:
            for folder_path, folder_info in folder_structure.items():
                if folder_info['level'] == 0 and folder_info['name'] != '列出資料名稱.py':
                    write_folder_structure(folder_info, file=text_file)
    
        # 創建Excel
        excel_file_path = os.path.join(current_directory, '資料階層及檔案清單.xlsx')
        workbook = Workbook()
        workbook.remove(workbook.active)  # 刪除default的工作表
    
        # 創建"All Folders"工作表，用於列出文件夾階層結構
        all_folders_sheet = workbook.create_sheet(title='All Folders')
        
        # 寫入檔夾階層結構和檔案名
        row = 1
        for folder_path, folder_info in folder_structure.items():
            folder_name = folder_info['name']
            level = folder_info['level']
            if folder_name not in exclude_list:
                folder_name_cell = all_folders_sheet.cell(row=row, column=level + 1)
                folder_name_cell.value = folder_name
                folder_sheet_name = re.sub(r'[\/:*?\\[\]]', '_', folder_name)[:30]
                folder_name_cell.hyperlink = f"資料階層及檔案清單.xlsx#'{folder_sheet_name}'!A1"
        
                # 獲取該檔夾內的檔案名
                files = folder_info['files']
                if files:
                    for file_name in files:
                        if file_name not in exclude_list:
                            file_name_cell = all_folders_sheet.cell(row=row + 1, column=max_folder_level+2)
                            file_path = os.path.join(folder_path, file_name)
                            file_hyperlink = f"{file_path[40:]}"
                            print(file_hyperlink)
                            file_name_cell.hyperlink = file_hyperlink
                            file_name_cell.value = file_name
                            row += 1
                row += 1
        
                folder_sheet = workbook.create_sheet(title=folder_sheet_name)
                folder_sheet['A1'] = folder_sheet_name
                folder_sheet.column_dimensions['A'].width = 50
                folder_sheet['A1'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
                subfolders = folder_info['subfolders']
                files = folder_info['files']
        
                if not subfolders and not files:
                    folder_sheet['A2'] = '找不到子檔夾或檔。'
                else:
                    subfolder_row = 2
                    for subfolder_path in subfolders:
                        subfolder_name = os.path.basename(subfolder_path)
                        subfolder_sheet_name = re.sub(r'[\/:*?\\[\]]', '_', subfolder_name)[:30]
                        folder_sheet.cell(row=subfolder_row, column=1).value = subfolder_name
                        folder_sheet.cell(row=subfolder_row, column=1).hyperlink = f"資料階層及檔案清單.xlsx#'{subfolder_sheet_name}'!A1"
                        subfolder_row += 1
        
                    file_row = subfolder_row
                    for file_name in files:
                        if file_name not in exclude_list:
                            file_path = os.path.join(folder_path, file_name)
                            file_hyperlink = f"{file_path[40:]}"
                            folder_sheet.cell(row=file_row, column=2).hyperlink = file_hyperlink
                            folder_sheet.cell(row=file_row, column=2).value = file_name
                            file_row += 1
        
            # 在每個工作表中添加超鏈接到第一張工作表
            if folder_sheet_name != 'All Folders':
                sheet = workbook[folder_sheet_name]
                sheet['A1'].hyperlink =  f"資料階層及檔案清單.xlsx#'All Folders'!{folder_name_cell.coordinate}"
                sheet['A1'].value = sheet['A1'].value + ' (返回總表)'
                
        # 調整總表的列寬
        all_folders_sheet.column_dimensions[get_column_letter(all_folders_sheet.max_column)].width = 30
    

        # 在每個工作表中添加超鏈接樣式
        hyperlink_style = Font(underline='single', color='0563C1')
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.hyperlink:
                        cell.font = hyperlink_style

        # 根據F欄的連續值 分組並收合
        def group_continuous_rows(sheet):
            max_row = sheet.max_row
            group_start = None
            previous_value = None
            for row in range(2, max_row + 1):
                current_value = sheet.cell(row=row, column=max_folder_level+1).value
        
                if current_value is not None:
                    if group_start is None:
                        group_start = row
                else:
                    if group_start is not None:
                        sheet.row_dimensions.group(group_start, row - 1, hidden=False)
                        group_start = None
        
        
        # 對"All Folders"工作表進行分組並收合
        group_continuous_rows(all_folders_sheet)
        
        # 對每個文件夾的工作表進行分組並收合
        for sheet_name in workbook.sheetnames:
            if sheet_name != 'All Folders':
                sheet = workbook[sheet_name]
                group_continuous_rows(sheet)

    
        # 保存Excel檔
        workbook.save(excel_file_path)
  


        # 在文字方塊中插入當前時間和文本內容
   
        text_output.insert(tk.END, f"資料夾、檔案的階層結構已匯出至【{current_directory}】:\n")
        text_output.insert(tk.END, f"【資料階層及檔案清單.txt】及【資料階層及檔案清單.xlsx】\n\n")
        # 獲取結束時間
        end_time = time.time()
        end_time_str = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(end_time))
        text_output.insert(tk.END, f"結束執行時間：{end_time_str}\n")
        # 計算總執行時間
        total_time = end_time - start_time
        text_output.insert(tk.END, f"程式執行時間：{total_time:.2f} 秒\n")  
        folder_path = current_directory  # 清空資料夾路徑

        # 計算行數
        line_count = text_output.get("1.0", "end-1c").count('\n') + 1
        
        # 設定text_output的高度
        text_output.config(height=line_count)


    except Exception as e:
        text_output.delete("1.0", tk.END)  # 清空原有內容
        # 獲取當前時間
        current_time = time.time()
        current_time = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(current_time))
        # 在文字方塊中插入當前時間和文本內容
        text_output.insert(tk.END, f"錯誤時間：{current_time}\n")
        text_output.insert(tk.END, f"錯誤: {str(e)}\n")
        folder_path = current_directory  # 清空資料夾路徑



# 資料夾路徑標籤
folder_label = ttk.Label(window, text="資料夾路徑:")
folder_label.grid(row=0, column=0, padx=10, pady=10)

# 選擇資料夾按鈕
select_button = ttk.Button(window, text="選擇資料夾", command=select_folder_path)
select_button.grid(row=0, column=1, padx=10, pady=10)

# 執行按鈕
execute_button = ttk.Button(window, text="執行", command=execute_program)
execute_button.grid(row=1, column=0, columnspan=2, padx=10, pady=10)

# 執行結果文字區域
text_output = tk.Text(window, height=10, width=60, wrap=tk.WORD)
text_output.grid(row=2, column=0, columnspan=2, padx=10, pady=10)



# 建立滾動條
scrollbar = ttk.Scrollbar(window, command=text_output.yview)
scrollbar.grid(row=2, column=2, sticky='ns')
text_output.config(yscrollcommand=scrollbar.set)

# 啟動主迴圈
window.mainloop()
